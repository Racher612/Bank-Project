#from Reading import reading_file, clear_transactions
from openpyxl import Workbook, styles
from openpyxl.styles import *



def create_file():
    pattern_list = ['Паттерн 1', "Паттерн 2", "Паттерн 3"]
    transactions = {'Паттерн 1' : ['1', '2', '3'], 'Паттерн 2' : ['4', '5', '6'], 'Паттерн 3' : ['7', '8', '9']}
    color_array = {'red' : 'FF0000', 'orange' : 'FF4500', 'yellow' : 'FFFF00', 'green' : '008000',
                   'blue' : '0000FF', 'purple' : '800080', 'white' : 'FFFFFF', 'grey' : 'DCDCDC',
                   'black' : '000000', 'crimson' : 'DC143C', 'aqua' : '00FFFF'}
    thins = Side(border_style="thin", color=color_array['black'])

    book = Workbook()
    frod = book.active

    #создание листа паттернов
    frod.title = "Фрод Паттерны"

    #создание шапки
    frod.column_dimensions['A'].width = 15.33203125
    frod.column_dimensions['B'].width = 28.58203125
    frod.row_dimensions[1].height = None
    frod.merge_cells('A1:B1')                   # cell is named A1
    frod['A1'].font = Font(color = "FF00B050", bold = True,
                                  name='Times New Roman', size=12)
    frod['A1'].border = Border(top=thins, left=thins, right=thins, bottom=thins)
    frod['A1'].fill = PatternFill("solid", fgColor = 'FFF2F2F2')
    frod['A1'] = "Выявленные фрод паттерны"
    frod['A1'].alignment = Alignment(horizontal="center", vertical="center")

    #задание шрифта таблицы
    for cell in frod['A2:B{}'.format(str(len(pattern_list) + 2))]:# + 2 потому что кол-во паттернов + шапка
        for cell2 in cell:
            cell2.font = Font(name='Times New Roman', size=12)
            cell2.alignment = Alignment(horizontal="center", vertical="center")

    # названия ячеек таблицы
    frod['A2'] = "№ паттерна"
    frod['A2'].border = Border(top=thins, left=thins, right=thins, bottom=thins)
    frod['B2'] = "Описание"
    frod['B2'].border = Border(top=thins, left=thins, right=thins, bottom=thins)

    #заполнение значений
    for i in range(len(pattern_list)):
        frod['A{}'.format(i + 3)] = i + 1
        frod['B{}'.format(i + 3)] = pattern_list[i]

    #создание листа транзакций
    trans = book.create_sheet("Транзакции")

    # создание шапки
    trans.column_dimensions['A'].width = 15.33203125
    trans.column_dimensions['B'].width = 50
    trans.row_dimensions[1].height = None
    trans.merge_cells('A1:B1')  # cell is named A1
    trans['A1'].font = Font(color="FF00B050", bold=True,
                           name='Times New Roman', size=12)
    trans['A1'].border = Border(top=thins, left=thins, right=thins, bottom=thins)
    trans['A1'].fill = PatternFill("solid", fgColor='FFF2F2F2')
    trans['A1'] = "Транзакции"
    trans['A1'].alignment = Alignment(horizontal="center", vertical="center")

    # задание шрифта таблицы
    for cell in trans['A2:B{}'.format(str(len(pattern_list) + 2))]:  # + 2 потому что кол-во паттернов + шапка
        for cell2 in cell:
            cell2.font = Font(name='Times New Roman', size=12)
            cell2.alignment = Alignment(horizontal="center", vertical="center")

    #создание наименований столбцов таблицы
    trans['A2'] = "№ паттерна"
    trans['A2'].border = Border(top=thins, left=thins, right=thins, bottom=thins)
    trans['B2'] = "Номера транзакций, попадающих под паттерн"
    trans['B2'].border = Border(top=thins, left=thins, right=thins, bottom=thins)

    # задание шрифта таблицы
    for i in range(len(pattern_list)):
        trans['A{}'.format(i + 3)] = i + 1
        trans['B{}'.format(i + 3)] = '[' + ', '.join(transactions[pattern_list[i]]) + ']'

    book.save("test 1.xlsx")

create_file()
